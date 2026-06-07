from __future__ import annotations

import copy
import json
import math
import os
import queue
import re
import shutil
import subprocess
import sys
import tempfile
import threading
import uuid
from pathlib import Path
from typing import Any

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import EASYcore as batch_core


APP_DIR = Path(__file__).resolve().parent
APP_VERSION = "version1"
APP_NAME = f"SNP Data Extractor Organizer_{APP_VERSION}"
APPDATA_DIR = Path(os.getenv("APPDATA") or Path.home() / "AppData" / "Roaming")
GUI_STATE_PATH = APPDATA_DIR / APP_NAME / "gui_state.json"
FILE_GROUPS = ("MAIN", "NEXT", "FEXT")
SUPPLIER_EXCEL_SHEETS = ("summary", "per_file", "standard deviation", "pure picture")
SUPPLIER_EXCEL_FONT = "Malgun Gothic Semilight"
SUPPLIER_SUMMARY_HEADER_MAP = {
    "data_group": "Group",
    "source_group": "Group",
    "parameter": "Para",
    "target_frequency_hz": "FREQs",
    "average_linear_magnitude_db": "linear-dB",
    "average_db": "AVG-dB",
    "minimum_db": "min dB",
    "maximum_db": "MAX dB",
    "files_used": "Files",
    "lower_frequency_hz_min": "Low Freq",
    "upper_frequency_hz_max": "High Freq",
}
SUPPLIER_STDEV_HEADER_MAP = {
    "data_group": "Group",
    "source_group": "Group",
    "parameter": "Para",
    "target_frequency_hz": "FREQs",
    "mean_linear_magnitude": "linear-dB",
    "standard_deviation_linear_magnitude": "linear-STD",
    "mean_db": "AVG-dB",
    "standard_deviation_db": "STD-dB",
    "minimum_db": "min dB",
    "maximum_db": "MAX dB",
    "sample_count": "Files",
}
SUPPLIER_PER_FILE_HEADER_MAP = {
    "data_group": "Group",
    "source_group": "Group",
    "parameter": "Para",
    "target_frequency_hz": "FREQs",
    "magnitude_db": "dB",
    "file_name": "File",
    "file_path": "Path",
    "result_frequency_hz": "Result Freq",
    "lower_frequency_hz": "Low Freq",
    "upper_frequency_hz": "High Freq",
    "linear_magnitude": "Linear",
    "interpolation_method": "Method",
}
RETURN_LOSS_ROW = "Return Loss (Sdd11 / Sdd22)"
SUPPLIER_DEFAULT_FREQUENCIES = {
    "Sdd21": "100MHz, 125MHz",
    RETURN_LOSS_ROW: "40MHz, 100MHz, 125MHz",
    "NEXT (Sdd21)": "125MHz",
    "FEXT (Sdd21)": "125MHz",
    "Scc21": "125MHz, 200MHz",
    "Scd22": "125MHz",
    "Sdc22": "125MHz",
}
SUPPLIER_DEFAULT_INCLUDED_PARAMETERS = {
    "Sdd21",
    "Sdd11",
    "Sdd22",
    "NEXT",
    "FEXT",
    "Scc21",
    "Scd22",
    "Sdc22",
}
SUPPLIER_DEFAULT_PLOT_LIMITS = {
    "x_frequency": {"min": "1MHz", "max": "500MHz"},
    "y_db": {"min": -50, "max": 0},
}
SUPPLIER_PLOT_Y_LIMITS = {
    "Sdd21": {"min": -5, "max": 0},
}
PARAMETER_ROWS = (
    ("Sdd21", ("Sdd21",)),
    (RETURN_LOSS_ROW, ("Sdd11", "Sdd22")),
    ("NEXT (Sdd21)", ("NEXT",)),
    ("FEXT (Sdd21)", ("FEXT",)),
    ("Scc21", ("Scc21",)),
    ("Scd22", ("Scd22",)),
    ("Sdc22", ("Sdc22",)),
)
PLOT_KEY_BY_LABEL = {
    "Sdd21": "Sdd21",
    RETURN_LOSS_ROW: "Sdd11_Sdd22",
    "NEXT (Sdd21)": "NEXT_Sdd21",
    "FEXT (Sdd21)": "FEXT_Sdd21",
    "Scc21": "Scc21",
    "Scd22": "Scd22",
    "Sdc22": "Sdc22",
}
THEME = {
    "app_bg": "#f6ead4",
    "card_bg": "#fff8ea",
    "card_border": "#d9bd8b",
    "ocean": "#0b7285",
    "ocean_dark": "#075461",
    "lagoon": "#d8f3ef",
    "sand": "#fff2cc",
    "coral": "#ff8a65",
    "coral_dark": "#d85c3c",
    "leaf": "#2f9e44",
    "text": "#17343a",
    "muted": "#5f6f72",
    "entry_bg": "#fffdf7",
    "selection": "#14a3a8",
}
DEFAULT_DOCUMENTS_BROWSE_DIR = Path.home() / "Documents"


def state_path_candidate(value: Any) -> Path | None:
    if not isinstance(value, str) or not value.strip():
        return None
    path = Path(value)
    return path if path.exists() else None


def existing_initial_dir(*candidates: Path | list[Path | None] | tuple[Path | None, ...] | None) -> Path:
    if len(candidates) == 1 and isinstance(candidates[0], (list, tuple)):
        candidates = tuple(candidates[0])
    for candidate in candidates:
        if candidate is None:
            continue
        if candidate.is_file():
            return candidate.parent
        if candidate.is_dir():
            return candidate
    return DEFAULT_DOCUMENTS_BROWSE_DIR if DEFAULT_DOCUMENTS_BROWSE_DIR.exists() else APP_DIR


def choose_folder_dialog(title: str, initial_dir: Path) -> str:
    return filedialog.askdirectory(title=title, initialdir=str(initial_dir))


def parse_optional_number(text: str, field_name: str) -> float | None:
    value = text.strip()
    if not value:
        return None
    try:
        number = float(value)
    except ValueError as exc:
        raise ValueError(f"{field_name} must be a number.") from exc
    if not math.isfinite(number):
        raise ValueError(f"{field_name} must be finite.")
    return number


def build_plot_limits_from_text(
    x_min: str,
    x_max: str,
    y_min: str,
    y_max: str,
) -> dict[str, Any]:
    x_min_text = x_min.strip() or None
    x_max_text = x_max.strip() or None
    y_min_value = parse_optional_number(y_min, "Y min dB")
    y_max_value = parse_optional_number(y_max, "Y max dB")
    x_min_hz = batch_core.parse_frequency(x_min_text) if x_min_text else None
    x_max_hz = batch_core.parse_frequency(x_max_text) if x_max_text else None
    if x_min_hz is not None and x_max_hz is not None and x_min_hz >= x_max_hz:
        raise ValueError("X min frequency must be smaller than X max frequency.")
    if y_min_value is not None and y_max_value is not None and y_min_value >= y_max_value:
        raise ValueError("Y min dB must be smaller than Y max dB.")
    return {
        "x_frequency": {"min": x_min_text, "max": x_max_text},
        "y_db": {"min": y_min_value, "max": y_max_value},
    }


def default_additional_plot_limits() -> dict[str, Any]:
    return {
        "x_frequency": {"min": "1MHz", "max": "500MHz"},
        "y_db": {"min": -50, "max": 0},
    }


def merge_plot_limit_defaults(
    current: dict[str, Any] | None,
    defaults: dict[str, Any],
) -> dict[str, Any]:
    merged = copy.deepcopy(defaults)
    if not isinstance(current, dict):
        return merged
    for section in ("x_frequency", "y_db"):
        if isinstance(current.get(section), dict):
            for key, value in current[section].items():
                if value is not None and str(value).strip():
                    merged.setdefault(section, {})[key] = value
    return merged


def validate_output_file_name(text: str) -> str:
    value = text.strip()
    if not value:
        raise ValueError("Output file name is required.")
    invalid = set('\\/:*?"<>|')
    if any(character in invalid for character in value):
        raise ValueError('Output file name cannot contain: \\ / : * ? " < > |')
    if value.endswith(".") or value.endswith(" "):
        raise ValueError("Output file name cannot end with a dot or space.")
    return value


def read_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as file:
        data = json.load(file)
    if not isinstance(data, dict):
        raise ValueError(f"JSON root must be an object: {path}")
    return data


def write_json(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as file:
        json.dump(data, file, ensure_ascii=False, indent=2)
        file.write("\n")


def read_gui_state() -> dict[str, Any]:
    if not GUI_STATE_PATH.exists():
        return {}
    try:
        return read_json(GUI_STATE_PATH)
    except Exception:
        return {}


def write_gui_state(data: dict[str, Any]) -> None:
    try:
        write_json(GUI_STATE_PATH, data)
    except OSError:
        pass


def current_entry_dir(raw_path: str) -> Path | None:
    text = raw_path.strip()
    if not text:
        return None
    path = Path(text)
    if path.is_file():
        return path.parent
    if path.is_dir():
        return path
    parent = path.parent
    return parent if str(parent) not in {"", "."} else None


def split_frequency_text(text: str) -> list[str]:
    values = [value.strip() for value in re.split(r"[,;\s]+", text.strip()) if value.strip()]
    for value in values:
        batch_core.parse_frequency(value)
    return sorted(set(values), key=batch_core.parse_frequency)


def default_frequency_texts() -> dict[str, str]:
    return dict(SUPPLIER_DEFAULT_FREQUENCIES)


def selected_frequency_map(
    parameter_settings: dict[str, tuple[bool, str]]
) -> tuple[dict[str, list[str]], list[str]]:
    frequency_map: dict[str, list[str]] = {}
    enabled_parameters: list[str] = []
    for label, parameters in PARAMETER_ROWS:
        include, frequency_text = parameter_settings[label]
        if not include:
            continue
        frequencies = split_frequency_text(frequency_text)
        if not frequencies:
            raise ValueError(f"{label} frequency list is empty.")
        for parameter in parameters:
            frequency_map[parameter] = list(frequencies)
            enabled_parameters.append(parameter)
    if not enabled_parameters:
        raise ValueError("Choose at least one parameter.")
    return frequency_map, enabled_parameters


def build_supplier_config(
    input_dir: str,
    output_dir: str,
    output_name: str,
    custom_output_name: bool,
    parameter_settings: dict[str, tuple[bool, str]],
    plot_limit_overrides: dict[str, dict[str, Any]],
    additional_targets: list[dict[str, Any]],
    run_temp_dir: Path,
) -> dict[str, Any]:
    frequency_map, enabled_parameters = selected_frequency_map(parameter_settings)
    plot_limits = {"global": copy.deepcopy(SUPPLIER_DEFAULT_PLOT_LIMITS)}
    for plot_key, limits in plot_limit_overrides.items():
        plot_limits[plot_key] = copy.deepcopy(limits)
    return {
        "input_dir": input_dir,
        "output_dir": output_dir,
        "output_naming": {
            "part_number_prefix": output_name,
            "custom_output_name": custom_output_name,
        },
        "recursive_search": False,
        "exclude_keywords": ["Golden"],
        "title_from_filename": {"enabled": False},
        "plot_legend": {"mode": "off"},
        "plot_average": {
            "direct_db": {
                "enabled": False,
                "color": "#ff0040",
                "linewidth": 3.0,
                "label": "Average dB",
            },
            "linear_magnitude_db": {
                "enabled": False,
                "color": "#00b7ff",
                "linewidth": 3.0,
                "label": "Average linear magnitude to dB",
            },
        },
        "xtk_plots": {
            "enabled": True,
            "folder_name": "XTK",
            "parameter": "Sdd21",
            "plot_types": ["NEXT", "FEXT"],
        },
        "supplier_excel_frequencies": frequency_map,
        "supplier_enabled_parameters": enabled_parameters,
        "additional_excel_targets": copy.deepcopy(additional_targets),
        "performance": {
            "processing_workers": 4,
            "trim_frequency": {
                "enabled": True,
                "margin_ratio": 1.05,
                "max_hz": None,
            },
        },
        "plot_limits": plot_limits,
    }


def trim_supplier_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    merge_supplier_return_loss_tables(workbook)
    append_supplier_additional_tables(workbook)
    rebuild_supplier_standard_deviation_sheet(workbook)
    move_additional_pictures_to_pure_sheet(workbook)
    format_supplier_result_sheets(workbook)
    keep = set(SUPPLIER_EXCEL_SHEETS)
    for sheet_name in list(workbook.sheetnames):
        if sheet_name not in keep:
            workbook.remove(workbook[sheet_name])
    ordered_sheets = [
        workbook[sheet_name]
        for sheet_name in SUPPLIER_EXCEL_SHEETS
        if sheet_name in workbook.sheetnames
    ]
    if ordered_sheets:
        workbook._sheets = ordered_sheets
    workbook.save(path)


def format_supplier_result_sheets(workbook: Any) -> None:
    sheet_settings = {
        "summary": SUPPLIER_SUMMARY_HEADER_MAP,
        "standard deviation": SUPPLIER_STDEV_HEADER_MAP,
        "per_file": SUPPLIER_PER_FILE_HEADER_MAP,
    }
    for sheet_name, header_map in sheet_settings.items():
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        header_row = find_result_header_row(sheet)
        if header_row is None:
            header_row = 1
        format_supplier_table_sheet(
            sheet,
            header_row,
            header_map,
            keep_only_mapped=sheet_name in {"summary", "standard deviation"},
        )


def find_result_header_row(worksheet: Any) -> int | None:
    for required_headers in (
        {"data_group", "parameter", "target_frequency_hz"},
        {"source_group", "parameter", "target_frequency_mhz"},
        {"file_name", "parameter", "target_frequency_hz"},
        {"parameter", "target_frequency_hz"},
    ):
        header_row = find_header_row(worksheet, required_headers)
        if header_row is not None:
            return header_row
    return None


def normalize_frequency_columns(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized = []
    for record in records:
        current = dict(record)
        for column in (
            "target_frequency_hz",
            "result_frequency_hz",
            "lower_frequency_hz_min",
            "upper_frequency_hz_max",
            "lower_frequency_hz",
            "upper_frequency_hz",
        ):
            value = safe_float(current.get(column))
            if value is not None:
                current[column] = value / 1e6
        normalized.append(current)
    return normalized


def supplier_sort_key(record: dict[str, Any]) -> tuple[str, str, float, str]:
    group = str(source_value(record) or "")
    parameter = str(record.get("parameter", ""))
    frequency = safe_float(record.get("target_frequency_hz")) or safe_float(record.get("target_frequency_mhz")) or 0.0
    file_name = str(record.get("file_name", ""))
    return (group, parameter, frequency, file_name)


def format_supplier_table_sheet(
    worksheet: Any,
    header_row: int,
    header_map: dict[str, str],
    keep_only_mapped: bool = False,
) -> None:
    headers, records = worksheet_records(worksheet, header_row)
    if keep_only_mapped:
        headers = [header for header in header_map if header in headers]
    else:
        preferred_headers = [header for header in header_map if header in headers]
        headers = [
            *preferred_headers,
            *[header for header in headers if header not in preferred_headers],
        ]
    if records:
        records = sorted(records, key=supplier_sort_key)
        records = normalize_frequency_columns(records)
        rewrite_records(worksheet, header_row, headers, records)
    if worksheet.max_column > len(headers):
        worksheet.delete_cols(len(headers) + 1, worksheet.max_column - len(headers))

    for column_index, original_header in enumerate(headers, start=1):
        worksheet.cell(header_row, column_index).value = header_map.get(
            original_header,
            original_header,
        )

    font = Font(name=SUPPLIER_EXCEL_FONT, size=10)
    header_font = Font(name=SUPPLIER_EXCEL_FONT, size=10, bold=True, color="17343A")
    header_fill = PatternFill("solid", fgColor="D8F3EF")
    group_fill_a = PatternFill("solid", fgColor="FFF8EA")
    group_fill_b = PatternFill("solid", fgColor="F3FBF8")
    thin_side = Side(style="thin", color="D9BD8B")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    group_column_index = None
    for column_index, original_header in enumerate(headers, start=1):
        if original_header in {"data_group", "source_group"}:
            group_column_index = column_index
            break

    current_group = object()
    fill_toggle = False
    for row_index in range(header_row, worksheet.max_row + 1):
        if row_index == header_row:
            row_fill = header_fill
            row_font = header_font
        else:
            if group_column_index is not None:
                group_value = worksheet.cell(row_index, group_column_index).value
                if group_value != current_group:
                    current_group = group_value
                    fill_toggle = not fill_toggle
            row_fill = group_fill_a if fill_toggle else group_fill_b
            row_font = font
        for column_index in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row_index, column_index)
            cell.font = row_font
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if row_index > header_row and isinstance(cell.value, float):
                cell.number_format = "0.0000"

    width_by_header = {
        "Group": 12,
        "Para": 18,
        "FREQs": 12,
        "linear-dB": 14,
        "AVG-dB": 12,
        "min dB": 12,
        "MAX dB": 12,
        "Files": 10,
        "Low Freq": 12,
        "High Freq": 12,
        "File": 28,
        "Path": 48,
    }
    for column_index in range(1, worksheet.max_column + 1):
        header_value = str(worksheet.cell(header_row, column_index).value or "")
        column_letter = worksheet.cell(header_row, column_index).column_letter
        worksheet.column_dimensions[column_letter].width = width_by_header.get(header_value, 16)

    worksheet.freeze_panes = worksheet.cell(header_row + 1, 1)
    worksheet.auto_filter.ref = worksheet.dimensions


def image_title_for_sheet(worksheet: Any, image: Any, fallback: str) -> str:
    anchor = getattr(image, "anchor", None)
    marker = getattr(anchor, "_from", None)
    if marker is None:
        return fallback
    title_row = max(1, int(marker.row))
    title_column = int(marker.col) + 1
    value = worksheet.cell(title_row, title_column).value
    return str(value).strip() if value else fallback


def move_additional_pictures_to_pure_sheet(workbook: Any) -> None:
    if "additional picture" not in workbook.sheetnames:
        return
    additional_sheet = workbook["additional picture"]
    additional_images = list(getattr(additional_sheet, "_images", []))
    if not additional_images:
        return

    if "pure picture" not in workbook.sheetnames:
        pure_sheet = workbook.create_sheet("pure picture")
    else:
        pure_sheet = workbook["pure picture"]

    anchor_columns = ("A", "H", "O")
    rows_per_image = 18
    start_index = len(getattr(pure_sheet, "_images", []))
    for offset, image in enumerate(additional_images):
        image_index = start_index + offset
        grid_column = image_index % len(anchor_columns)
        grid_row = image_index // len(anchor_columns)
        current_row = grid_row * rows_per_image + 1
        column = anchor_columns[grid_column]
        title = image_title_for_sheet(additional_sheet, image, f"Additional_{offset + 1}")
        pure_sheet[f"{column}{current_row}"] = title
        pure_sheet[f"{column}{current_row}"].style = "Title"
        image.anchor = f"{column}{current_row + 1}"
        pure_sheet.add_image(image)

    additional_sheet._images = []


def source_value(record: dict[str, Any]) -> Any:
    return record.get("data_group", record.get("source_group"))


def hz_from_record(record: dict[str, Any]) -> Any:
    if record.get("target_frequency_hz") is not None:
        return record.get("target_frequency_hz")
    mhz_value = safe_float(record.get("target_frequency_mhz"))
    return None if mhz_value is None else mhz_value * 1e6


def append_records_to_existing_table(
    worksheet: Any,
    header_row: int,
    target_headers: list[str],
    source_records: list[dict[str, Any]],
) -> None:
    start_row = worksheet.max_row + 1
    for row_offset, record in enumerate(source_records):
        row_index = start_row + row_offset
        for column_index, header in enumerate(target_headers, start=1):
            if header == "data_group":
                value = source_value(record)
            elif header == "target_frequency_hz":
                value = hz_from_record(record)
            elif header == "lower_frequency_hz_min":
                value = record.get("lower_frequency_hz_min", record.get("lower_frequency_hz"))
            elif header == "upper_frequency_hz_max":
                value = record.get("upper_frequency_hz_max", record.get("upper_frequency_hz"))
            else:
                value = record.get(header)
            worksheet.cell(row_index, column_index).value = value


def append_supplier_additional_tables(workbook: Any) -> None:
    if "summary" in workbook.sheetnames and "additional results" in workbook.sheetnames:
        summary_sheet = workbook["summary"]
        header_row = find_header_row(
            summary_sheet,
            {"data_group", "parameter", "target_frequency_hz"},
        )
        if header_row is not None:
            summary_headers, _ = worksheet_records(summary_sheet, header_row)
            _, additional_records = worksheet_records(workbook["additional results"], 1)
            append_records_to_existing_table(
                summary_sheet,
                header_row,
                summary_headers,
                additional_records,
            )

    if "per_file" in workbook.sheetnames and "additional per file" in workbook.sheetnames:
        per_file_sheet = workbook["per_file"]
        header_row = find_header_row(
            per_file_sheet,
            {"file_name", "parameter", "target_frequency_hz", "magnitude_db"},
        )
        if header_row is not None:
            per_file_headers, _ = worksheet_records(per_file_sheet, header_row)
            _, additional_records = worksheet_records(workbook["additional per file"], 1)
            append_records_to_existing_table(
                per_file_sheet,
                header_row,
                per_file_headers,
                additional_records,
            )


def find_header_row(worksheet: Any, required_headers: set[str]) -> int | None:
    for row_index in range(1, worksheet.max_row + 1):
        values = {
            str(cell.value).strip()
            for cell in worksheet[row_index]
            if cell.value is not None
        }
        if required_headers.issubset(values):
            return row_index
    return None


def worksheet_records(worksheet: Any, header_row: int) -> tuple[list[str], list[dict[str, Any]]]:
    headers = [
        str(worksheet.cell(header_row, column_index).value).strip()
        for column_index in range(1, worksheet.max_column + 1)
        if worksheet.cell(header_row, column_index).value is not None
    ]
    records: list[dict[str, Any]] = []
    for row_index in range(header_row + 1, worksheet.max_row + 1):
        values = {
            header: worksheet.cell(row_index, column_index).value
            for column_index, header in enumerate(headers, start=1)
        }
        if any(value is not None for value in values.values()):
            records.append(values)
    return headers, records


def rewrite_records(worksheet: Any, header_row: int, headers: list[str], records: list[dict[str, Any]]) -> None:
    if worksheet.max_row > header_row:
        worksheet.delete_rows(header_row + 1, worksheet.max_row - header_row)
    for row_offset, record in enumerate(records, start=1):
        row_index = header_row + row_offset
        for column_index, header in enumerate(headers, start=1):
            worksheet.cell(row_index, column_index).value = record.get(header)


def sample_standard_deviation(values: list[float]) -> float | None:
    if len(values) < 2:
        return None
    mean_value = sum(values) / len(values)
    variance = sum((value - mean_value) ** 2 for value in values) / (len(values) - 1)
    return math.sqrt(variance)


def rebuild_supplier_standard_deviation_sheet(workbook: Any) -> None:
    if "per_file" not in workbook.sheetnames:
        return

    per_file_sheet = workbook["per_file"]
    header_row = find_header_row(
        per_file_sheet,
        {"file_name", "parameter", "target_frequency_hz", "magnitude_db"},
    )
    if header_row is None:
        return

    _headers, per_file_records = worksheet_records(per_file_sheet, header_row)
    grouped: dict[tuple[Any, str, float], list[dict[str, Any]]] = {}
    for record in per_file_records:
        parameter = str(record.get("parameter", "")).strip()
        target_hz = hz_from_record(record)
        target_hz_value = safe_float(target_hz)
        if not parameter or target_hz_value is None:
            continue
        data_group = source_value(record) or "MAIN"
        grouped.setdefault((data_group, parameter, target_hz_value), []).append(record)

    output_headers = [
        "data_group",
        "parameter",
        "target_frequency_hz",
        "target_frequency_mhz",
        "sample_count",
        "mean_db",
        "standard_deviation_db",
        "mean_linear_magnitude",
        "standard_deviation_linear_magnitude",
        "minimum_db",
        "maximum_db",
    ]
    output_records: list[dict[str, Any]] = []
    for (data_group, parameter, target_hz), records in sorted(
        grouped.items(),
        key=lambda item: (str(item[0][0]), str(item[0][1]), float(item[0][2])),
    ):
        db_values = [
            value
            for value in (safe_float(record.get("magnitude_db")) for record in records)
            if value is not None
        ]
        linear_values = [
            value
            for value in (safe_float(record.get("linear_magnitude")) for record in records)
            if value is not None
        ]
        if not db_values:
            continue
        output_records.append(
            {
                "data_group": data_group,
                "parameter": parameter,
                "target_frequency_hz": target_hz,
                "target_frequency_mhz": target_hz / 1e6,
                "sample_count": len(db_values),
                "mean_db": sum(db_values) / len(db_values),
                "standard_deviation_db": sample_standard_deviation(db_values),
                "mean_linear_magnitude": (
                    sum(linear_values) / len(linear_values) if linear_values else None
                ),
                "standard_deviation_linear_magnitude": sample_standard_deviation(linear_values),
                "minimum_db": min(db_values),
                "maximum_db": max(db_values),
            }
        )

    if "standard deviation" in workbook.sheetnames:
        sheet = workbook["standard deviation"]
        if sheet.max_row:
            sheet.delete_rows(1, sheet.max_row)
    else:
        sheet = workbook.create_sheet("standard deviation")

    for column_index, header in enumerate(output_headers, start=1):
        sheet.cell(1, column_index).value = header
    for row_index, record in enumerate(output_records, start=2):
        for column_index, header in enumerate(output_headers, start=1):
            sheet.cell(row_index, column_index).value = record.get(header)


def safe_float(value: Any) -> float | None:
    try:
        if value is None:
            return None
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def db_to_linear(db_value: float) -> float:
    return 10 ** (db_value / 20)


def linear_to_db(linear_value: float) -> float | None:
    if linear_value <= 0:
        return None
    return 20 * math.log10(linear_value)


def weighted_average(values: list[float], weights: list[float]) -> float | None:
    total_weight = sum(weights)
    if not values or total_weight <= 0:
        return None
    return sum(value * weight for value, weight in zip(values, weights)) / total_weight


def combine_summary_return_loss(rows: list[dict[str, Any]]) -> dict[str, Any]:
    base = dict(rows[0])
    base["parameter"] = "Return Loss"

    weights = [
        safe_float(row.get("files_used")) or 1.0
        for row in rows
    ]
    average_values = [
        value
        for value in (safe_float(row.get("average_db")) for row in rows)
        if value is not None
    ]
    if len(average_values) == len(rows):
        base["average_db"] = weighted_average(average_values, weights)

    linear_average_values = [
        value
        for value in (safe_float(row.get("average_linear_magnitude_db")) for row in rows)
        if value is not None
    ]
    if len(linear_average_values) == len(rows):
        linear_average = weighted_average(
            [db_to_linear(value) for value in linear_average_values],
            weights,
        )
        if linear_average is not None:
            base["average_linear_magnitude_db"] = linear_to_db(linear_average)

    maximum_values = [
        value
        for value in (safe_float(row.get("maximum_db")) for row in rows)
        if value is not None
    ]
    if maximum_values:
        base["maximum_db"] = max(maximum_values)

    minimum_values = [
        value
        for value in (safe_float(row.get("minimum_db")) for row in rows)
        if value is not None
    ]
    if minimum_values:
        base["minimum_db"] = min(minimum_values)

    file_counts = [
        value
        for value in (safe_float(row.get("files_used")) for row in rows)
        if value is not None
    ]
    if file_counts:
        base["files_used"] = max(file_counts)

    lower_values = [
        value
        for value in (safe_float(row.get("lower_frequency_hz_min")) for row in rows)
        if value is not None
    ]
    if lower_values:
        base["lower_frequency_hz_min"] = min(lower_values)

    upper_values = [
        value
        for value in (safe_float(row.get("upper_frequency_hz_max")) for row in rows)
        if value is not None
    ]
    if upper_values:
        base["upper_frequency_hz_max"] = max(upper_values)

    return base


def combine_per_file_return_loss(rows: list[dict[str, Any]]) -> dict[str, Any]:
    base = dict(rows[0])
    base["parameter"] = "Return Loss"

    linear_values = [
        value
        for value in (safe_float(row.get("linear_magnitude")) for row in rows)
        if value is not None
    ]
    if linear_values:
        base["linear_magnitude"] = sum(linear_values) / len(linear_values)

    db_values = [
        value
        for value in (safe_float(row.get("magnitude_db")) for row in rows)
        if value is not None
    ]
    if db_values:
        base["magnitude_db"] = sum(db_values) / len(db_values)

    return base


def merge_return_loss_records(
    records: list[dict[str, Any]],
    summary_table: bool,
) -> list[dict[str, Any]]:
    if not records:
        return records

    excluded_columns = {
        "parameter",
        "average_db",
        "average_linear_magnitude_db",
        "maximum_db",
        "minimum_db",
        "files_used",
        "lower_frequency_hz_min",
        "upper_frequency_hz_max",
    } if summary_table else {"parameter", "linear_magnitude", "magnitude_db"}

    grouped_rows: dict[tuple[tuple[str, Any], ...], list[dict[str, Any]]] = {}
    for record in records:
        if str(record.get("parameter", "")).strip() not in {"Sdd11", "Sdd22"}:
            continue
        group_key = tuple(
            (column, record.get(column))
            for column in record
            if column not in excluded_columns
        )
        grouped_rows.setdefault(group_key, []).append(record)

    combined_by_key: dict[tuple[tuple[str, Any], ...], dict[str, Any]] = {}
    for group_key, rows in grouped_rows.items():
        present_parameters = {str(row.get("parameter", "")).strip() for row in rows}
        if {"Sdd11", "Sdd22"}.issubset(present_parameters):
            combiner = combine_summary_return_loss if summary_table else combine_per_file_return_loss
            combined_by_key[group_key] = combiner(rows)

    emitted: set[tuple[tuple[str, Any], ...]] = set()
    merged_records: list[dict[str, Any]] = []
    for record in records:
        parameter = str(record.get("parameter", "")).strip()
        if parameter not in {"Sdd11", "Sdd22"}:
            merged_records.append(record)
            continue
        group_key = tuple(
            (column, record.get(column))
            for column in record
            if column not in excluded_columns
        )
        combined = combined_by_key.get(group_key)
        if combined is None:
            merged_records.append(record)
            continue
        if group_key not in emitted:
            merged_records.append(combined)
            emitted.add(group_key)
    return merged_records


def merge_supplier_return_loss_tables(workbook: Any) -> None:
    if "summary" in workbook.sheetnames:
        summary_sheet = workbook["summary"]
        header_row = find_header_row(
            summary_sheet,
            {"data_group", "parameter", "target_frequency_hz"},
        )
        if header_row is not None:
            headers, records = worksheet_records(summary_sheet, header_row)
            rewrite_records(
                summary_sheet,
                header_row,
                headers,
                merge_return_loss_records(records, summary_table=True),
            )

    if "per_file" in workbook.sheetnames:
        per_file_sheet = workbook["per_file"]
        header_row = find_header_row(
            per_file_sheet,
            {"file_name", "parameter", "target_frequency_hz", "magnitude_db"},
        )
        if header_row is not None:
            headers, records = worksheet_records(per_file_sheet, header_row)
            rewrite_records(
                per_file_sheet,
                header_row,
                headers,
                merge_return_loss_records(records, summary_table=False),
            )


class SupplierSimpleGui(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title(f"SNP Data Extractor & Organizer - {APP_VERSION}")
        self.geometry("1200x780")
        self.minsize(960, 560)

        self.gui_state = read_gui_state()
        self.output_dir = tk.StringVar(value=str(self.gui_state.get("last_output_dir", "")))
        self.output_name = tk.StringVar()
        self.custom_output_name = tk.BooleanVar(value=False)
        self.selected_group = tk.StringVar(value="MAIN")
        self.output_status = tk.StringVar(value="Add MAIN/NEXT/FEXT .s4p files.")
        self.file_count_status = tk.StringVar(value="Selected files: MAIN 0, NEXT 0, FEXT 0.")
        self.progress_value = tk.IntVar(value=0)
        self.status = tk.StringVar(value="Ready")
        self.progress_label = tk.StringVar(value="0%")
        self.process: subprocess.Popen[str] | None = None
        self.process_queue: queue.Queue[str] = queue.Queue()
        self.completed_excel_path: Path | None = None
        self.selected_files: dict[str, dict[str, str]] = {}
        self.run_temp_dir: Path | None = None
        self.parameter_include_vars: dict[str, tk.BooleanVar] = {}
        self.parameter_frequency_vars: dict[str, tk.StringVar] = {}
        self.parameter_plot_limits: dict[str, dict[str, Any]] = {}
        self.additional_rows: dict[str, dict[str, Any]] = {}
        self.default_parameter_frequencies = default_frequency_texts()
        self._loading = False

        self._configure_theme()
        self._build_ui()
        self.after(120, self._poll_process_queue)

    def _configure_theme(self) -> None:
        self.configure(bg=THEME["app_bg"])
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass

        default_font = ("Segoe UI", 10)
        heading_font = ("Segoe UI", 10, "bold")
        style.configure(".", font=default_font)
        style.configure("App.TFrame", background=THEME["app_bg"])
        style.configure("Card.TFrame", background=THEME["card_bg"])
        style.configure(
            "Card.TLabelframe",
            background=THEME["card_bg"],
            bordercolor=THEME["card_border"],
            relief="solid",
            padding=10,
        )
        style.configure(
            "Card.TLabelframe.Label",
            background=THEME["card_bg"],
            foreground=THEME["ocean_dark"],
            font=("Segoe UI", 10, "bold"),
        )
        style.configure("TLabel", background=THEME["app_bg"], foreground=THEME["text"])
        style.configure("Card.TLabel", background=THEME["card_bg"], foreground=THEME["text"])
        style.configure("Muted.TLabel", background=THEME["card_bg"], foreground=THEME["muted"])
        style.configure("Hint.TLabel", background=THEME["lagoon"], foreground=THEME["ocean_dark"])
        style.configure("TCheckbutton", background=THEME["app_bg"], foreground=THEME["text"])
        style.configure("Card.TCheckbutton", background=THEME["card_bg"], foreground=THEME["text"])
        style.configure(
            "TNotebook",
            background=THEME["app_bg"],
            borderwidth=0,
            tabmargins=(4, 4, 4, 0),
        )
        style.configure(
            "TNotebook.Tab",
            background="#ead6ad",
            foreground=THEME["text"],
            padding=(14, 7),
        )
        style.map(
            "TNotebook.Tab",
            background=[("selected", THEME["card_bg"])],
            foreground=[("selected", THEME["ocean_dark"])],
        )
        style.configure(
            "TButton",
            background=THEME["lagoon"],
            foreground=THEME["ocean_dark"],
            bordercolor=THEME["ocean"],
            padding=(10, 5),
        )
        style.map(
            "TButton",
            background=[("active", "#c2ebe4"), ("disabled", "#d7d0c3")],
            foreground=[("disabled", "#8a8275")],
        )
        style.configure(
            "Primary.TButton",
            background=THEME["coral"],
            foreground="#ffffff",
            bordercolor=THEME["coral_dark"],
            padding=(12, 6),
            font=("Segoe UI", 10, "bold"),
        )
        style.map("Primary.TButton", background=[("active", "#ff7043")])
        style.configure(
            "Tool.TButton",
            background=THEME["sand"],
            foreground=THEME["ocean_dark"],
            bordercolor=THEME["card_border"],
            padding=(9, 4),
        )
        style.configure(
            "TEntry",
            fieldbackground=THEME["entry_bg"],
            background=THEME["entry_bg"],
            foreground=THEME["text"],
            bordercolor=THEME["card_border"],
        )
        style.configure(
            "TCombobox",
            fieldbackground=THEME["entry_bg"],
            background=THEME["entry_bg"],
            foreground=THEME["text"],
            bordercolor=THEME["card_border"],
            arrowcolor=THEME["ocean"],
        )
        style.configure(
            "Treeview",
            background="#fffef8",
            fieldbackground="#fffef8",
            foreground=THEME["text"],
            rowheight=25,
            bordercolor=THEME["card_border"],
        )
        style.configure(
            "Treeview.Heading",
            background=THEME["lagoon"],
            foreground=THEME["ocean_dark"],
            font=heading_font,
            relief="flat",
        )
        style.map(
            "Treeview",
            background=[("selected", THEME["selection"])],
            foreground=[("selected", "#ffffff")],
        )
        style.configure(
            "TProgressbar",
            background=THEME["ocean"],
            troughcolor="#f0d8ac",
            bordercolor=THEME["card_border"],
            lightcolor=THEME["ocean"],
            darkcolor=THEME["ocean_dark"],
        )

    def _build_ui(self) -> None:
        outer = ttk.Frame(self, padding=14, style="App.TFrame")
        outer.pack(fill="both", expand=True)

        banner = tk.Frame(
            outer,
            bg=THEME["ocean"],
            highlightthickness=0,
        )
        banner.pack(fill="x", pady=(0, 8))
        banner_inner = tk.Frame(banner, bg=THEME["ocean"])
        banner_inner.pack(fill="x", padx=18, pady=10)
        tk.Label(
            banner_inner,
            text=f"SNP Data Extractor & Organizer - {APP_VERSION}",
            bg=THEME["ocean"],
            fg="#ffffff",
            font=("Segoe UI", 18, "bold"),
        ).pack(anchor="w")
        tk.Label(
            banner_inner,
            text=(
                "4-port .s4p data extraction, mixed-mode charts, and Excel organization\n"
                "File support: only 4-port .s4p | Port map: Pair1=P1/P2, Pair2=P3/P4 | "
                "Reference impedance: single-ended / differential / common = 50 / 100 / 25 ohm\n"
                "檔案支援：僅支援 4-port .s4p | Port 定義：Pair1=P1/P2, Pair2=P3/P4 | "
                "參考阻抗：單端 / 差模 / 共模 = 50 / 100 / 25 ohm"
            ),
            bg=THEME["ocean"],
            fg="#d8f3ef",
            font=("Segoe UI", 10),
            justify="left",
            wraplength=1120,
        ).pack(anchor="w", pady=(3, 0), fill="x")

        notebook = ttk.Notebook(outer)
        notebook.pack(fill="both", expand=True)
        setup_tab = ttk.Frame(notebook, padding=8, style="App.TFrame")
        frequency_tab = ttk.Frame(notebook, padding=12, style="App.TFrame")
        notebook.add(setup_tab, text="Setup")
        notebook.add(frequency_tab, text="Frequencies & Parameters")

        files_frame = ttk.LabelFrame(setup_tab, text="S4P Files", padding=10, style="Card.TLabelframe")
        files_frame.pack(fill="x", pady=(0, 8))
        file_actions = ttk.Frame(files_frame, style="Card.TFrame")
        file_actions.pack(fill="x", pady=(0, 8))
        ttk.Label(file_actions, text="Group", style="Card.TLabel").pack(side="left", padx=(0, 8))
        ttk.Combobox(
            file_actions,
            textvariable=self.selected_group,
            values=FILE_GROUPS,
            state="readonly",
            width=10,
        ).pack(side="left", padx=(0, 14))
        ttk.Button(file_actions, text="Add .s4p Files", command=self.add_s4p_files).pack(
            side="left", padx=(0, 8)
        )
        ttk.Button(
            file_actions,
            text="Remove Selected",
            command=self.remove_selected_files,
        ).pack(side="left", padx=(0, 8))
        ttk.Button(file_actions, text="Clear", command=self.clear_selected_files).pack(side="left")

        file_holder = ttk.Frame(files_frame, style="Card.TFrame")
        file_holder.pack(fill="x")
        self.file_table = ttk.Treeview(
            file_holder,
            columns=("group", "file_name", "path"),
            show="headings",
            height=5,
        )
        self.file_table.heading("group", text="Group")
        self.file_table.heading("file_name", text="File name")
        self.file_table.heading("path", text="Original path")
        self.file_table.column("group", width=90, stretch=False, anchor="center")
        self.file_table.column("file_name", width=240, stretch=False)
        self.file_table.column("path", width=620)
        self.file_table.pack(side="left", fill="x", expand=True)
        file_scroll = ttk.Scrollbar(file_holder, orient="vertical", command=self.file_table.yview)
        file_scroll.pack(side="right", fill="y")
        self.file_table.configure(yscrollcommand=file_scroll.set)

        paths = ttk.LabelFrame(setup_tab, text="Output", padding=10, style="Card.TLabelframe")
        paths.pack(fill="x", pady=(0, 8))
        self._path_row(paths, "Output folder", self.output_dir, self.choose_output).pack(
            fill="x", pady=4
        )

        name_row = ttk.Frame(paths, style="Card.TFrame")
        name_row.pack(fill="x", pady=4)
        ttk.Label(name_row, text="Output file name", width=18, style="Card.TLabel").pack(side="left", padx=(0, 8))
        self.output_combo = ttk.Combobox(name_row, textvariable=self.output_name, width=34)
        self.output_combo.pack(side="left")
        ttk.Checkbutton(
            name_row,
            text="Custom",
            variable=self.custom_output_name,
            command=self._toggle_output_mode,
        ).pack(side="left", padx=(10, 0))

        status_row = ttk.Frame(paths, style="Card.TFrame")
        status_row.pack(fill="x", pady=(4, 0))
        ttk.Label(status_row, text="", width=18, style="Card.TLabel").pack(side="left", padx=(0, 8))
        ttk.Label(status_row, textvariable=self.output_status, wraplength=780, style="Muted.TLabel").pack(
            side="left", fill="x", expand=True
        )
        count_row = ttk.Frame(paths, style="Card.TFrame")
        count_row.pack(fill="x", pady=(2, 0))
        ttk.Label(count_row, text="", width=18, style="Card.TLabel").pack(side="left", padx=(0, 8))
        ttk.Label(count_row, textvariable=self.file_count_status, wraplength=780, style="Muted.TLabel").pack(
            side="left", fill="x", expand=True
        )

        actions = ttk.Frame(setup_tab, style="App.TFrame")
        actions.pack(fill="x", pady=(0, 10))
        self.start_button = ttk.Button(
            actions,
            text="Start Processing",
            command=self.start_processing,
            style="Primary.TButton",
        )
        self.start_button.pack(side="left")
        self.stop_button = ttk.Button(
            actions,
            text="Stop",
            command=self.stop_processing,
            state="disabled",
        )
        self.stop_button.pack(side="left", padx=(8, 0))
        ttk.Label(actions, textvariable=self.status).pack(side="left", padx=(18, 0))

        progress_row = ttk.Frame(setup_tab, style="App.TFrame")
        progress_row.pack(fill="x", pady=(0, 10))
        self.progress_bar = ttk.Progressbar(
            progress_row,
            orient="horizontal",
            mode="determinate",
            maximum=100,
            variable=self.progress_value,
        )
        self.progress_bar.pack(side="left", fill="x", expand=True)
        ttk.Label(progress_row, textvariable=self.progress_label).pack(side="left", padx=(8, 0))

        log_frame = ttk.LabelFrame(setup_tab, text="Processing Log", padding=8, style="Card.TLabelframe")
        log_frame.pack(fill="both", expand=True)
        self.log_text = tk.Text(
            log_frame,
            wrap="word",
            height=24,
            bg="#fffef8",
            fg=THEME["text"],
            insertbackground=THEME["ocean_dark"],
            relief="flat",
            highlightthickness=1,
            highlightbackground=THEME["card_border"],
        )
        self.log_text.pack(side="left", fill="both", expand=True)
        scroll = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        scroll.pack(side="right", fill="y")
        self.log_text.configure(yscrollcommand=scroll.set)

        self._toggle_output_mode()
        self._build_frequency_tab(frequency_tab)

    def _build_frequency_tab(self, parent: ttk.Frame) -> None:
        ttk.Label(
            parent,
            text=(
                "Choose which parameters are written to Excel and edit target "
                "frequencies. Examples: 100MHz, 125MHz, 200MHz"
            ),
            wraplength=880,
            style="Hint.TLabel",
        ).pack(fill="x", pady=(0, 10))

        table_card = ttk.LabelFrame(parent, text="Default Parameters", padding=12, style="Card.TLabelframe")
        table_card.pack(fill="x", anchor="n")
        table = ttk.Frame(table_card, style="Card.TFrame")
        table.pack(fill="x", anchor="n")
        ttk.Label(table, text="Include", width=10, style="Card.TLabel").grid(row=0, column=0, sticky="w", padx=(0, 8))
        ttk.Label(table, text="Parameter", width=28, style="Card.TLabel").grid(row=0, column=1, sticky="w", padx=(0, 8))
        ttk.Label(table, text="Frequency points", style="Card.TLabel").grid(row=0, column=2, sticky="w")
        ttk.Label(table, text="Plot", style="Card.TLabel").grid(row=0, column=3, sticky="w", padx=(8, 0))

        for row_index, (label, parameters) in enumerate(PARAMETER_ROWS, start=1):
            include_default = any(
                parameter in SUPPLIER_DEFAULT_INCLUDED_PARAMETERS for parameter in parameters
            )
            include_var = tk.BooleanVar(value=include_default)
            frequency_var = tk.StringVar(
                value=self.default_parameter_frequencies.get(label, "")
            )
            self.parameter_include_vars[label] = include_var
            self.parameter_frequency_vars[label] = frequency_var
            ttk.Checkbutton(table, variable=include_var, style="Card.TCheckbutton").grid(
                row=row_index, column=0, sticky="w", padx=(0, 8), pady=4
            )
            ttk.Label(table, text=label, width=28, style="Card.TLabel").grid(
                row=row_index, column=1, sticky="w", padx=(0, 8), pady=4
            )
            ttk.Entry(table, textvariable=frequency_var, width=70).grid(
                row=row_index, column=2, sticky="ew", pady=4
            )
            ttk.Button(
                table,
                text="Plot Setting",
                command=lambda current_label=label: self.edit_default_plot_setting(current_label),
                style="Tool.TButton",
            ).grid(row=row_index, column=3, sticky="w", padx=(8, 0), pady=4)
        table.columnconfigure(2, weight=1)

        buttons = ttk.Frame(parent, style="App.TFrame")
        buttons.pack(fill="x", pady=(12, 0))
        ttk.Button(buttons, text="Reset Defaults", command=self.reset_parameter_defaults, style="Tool.TButton").pack(
            side="left"
        )
        ttk.Button(buttons, text="Validate Frequencies", command=self.validate_frequencies_popup, style="Tool.TButton").pack(
            side="left", padx=(8, 0)
        )
        ttk.Button(buttons, text="Save Config", command=self.save_frequency_config, style="Primary.TButton").pack(
            side="left", padx=(18, 0)
        )
        ttk.Button(buttons, text="Recall Config", command=self.recall_frequency_config, style="Tool.TButton").pack(
            side="left", padx=(8, 0)
        )

        additional = ttk.LabelFrame(parent, text="Additional Parameters", padding=12, style="Card.TLabelframe")
        additional.pack(fill="both", expand=True, pady=(16, 0))
        additional_actions = ttk.Frame(additional, style="Card.TFrame")
        additional_actions.pack(fill="x", pady=(0, 8))
        ttk.Button(
            additional_actions,
            text="Add Parameter Row",
            command=self.add_additional_row,
            style="Tool.TButton",
        ).pack(side="left", padx=(0, 6))
        ttk.Button(
            additional_actions,
            text="Clear",
            command=self.clear_additional_rows,
            style="Tool.TButton",
        ).pack(side="left")

        header = ttk.Frame(additional, style="Card.TFrame")
        header.pack(fill="x", pady=(0, 2))
        ttk.Label(header, text="Source", width=12, style="Card.TLabel").grid(row=0, column=0, sticky="w", padx=(0, 8))
        ttk.Label(header, text="Parameter", width=16, style="Card.TLabel").grid(row=0, column=1, sticky="w", padx=(0, 8))
        ttk.Label(header, text="Frequency points", style="Card.TLabel").grid(row=0, column=2, sticky="w")
        ttk.Label(header, text="Plot", style="Card.TLabel").grid(row=0, column=3, sticky="w", padx=(8, 0))
        header.columnconfigure(2, weight=1)

        self.additional_rows_frame = ttk.Frame(additional, style="Card.TFrame")
        self.additional_rows_frame.pack(fill="both", expand=True)

    def _path_row(
        self,
        parent: ttk.Frame,
        label: str,
        variable: tk.StringVar,
        command,
    ) -> ttk.Frame:
        row = ttk.Frame(parent, style="Card.TFrame")
        ttk.Label(row, text=label, width=18, style="Card.TLabel").pack(side="left", padx=(0, 8))
        ttk.Entry(row, textvariable=variable).pack(side="left", fill="x", expand=True)
        ttk.Button(row, text="Browse", command=command, width=12, style="Tool.TButton").pack(side="left", padx=(8, 0))
        return row

    def parameter_settings(self) -> dict[str, tuple[bool, str]]:
        return {
            label: (
                bool(self.parameter_include_vars[label].get()),
                self.parameter_frequency_vars[label].get().strip(),
            )
            for label, _parameters in PARAMETER_ROWS
        }

    def reset_parameter_defaults(self) -> None:
        self.default_parameter_frequencies = default_frequency_texts()
        for label, parameters in PARAMETER_ROWS:
            self.parameter_include_vars[label].set(
                any(parameter in SUPPLIER_DEFAULT_INCLUDED_PARAMETERS for parameter in parameters)
            )
            self.parameter_frequency_vars[label].set(
                self.default_parameter_frequencies.get(label, "")
            )
        self.parameter_plot_limits.clear()

    def validate_frequencies_popup(self) -> None:
        try:
            selected_frequency_map(self.parameter_settings())
            self.additional_targets_for_config()
        except Exception as exc:
            messagebox.showerror("Frequency settings", str(exc), parent=self)
        else:
            messagebox.showinfo("Frequency settings", "Frequency settings are valid.", parent=self)

    def frequency_config_payload(self) -> dict[str, Any]:
        return {
            "version": f"easy-frequency-config-{APP_VERSION}",
            "default_parameters": {
                label: {
                    "include": bool(self.parameter_include_vars[label].get()),
                    "frequency_points": self.parameter_frequency_vars[label].get().strip(),
                    "plot_limits": copy.deepcopy(self.parameter_plot_limits.get(label)),
                }
                for label, _parameters in PARAMETER_ROWS
            },
            "additional_parameters": [
                {
                    "source": row["source_var"].get().strip().upper(),
                    "parameter": row["parameter_var"].get().strip(),
                    "frequency_points": row["frequency_var"].get().strip(),
                    "plot_limits": copy.deepcopy(row.get("plot_limits")),
                }
                for row in self.additional_rows.values()
            ],
        }

    def save_frequency_config(self) -> None:
        try:
            selected_frequency_map(self.parameter_settings())
            self.additional_targets_for_config()
        except Exception as exc:
            messagebox.showerror("Save Config", str(exc), parent=self)
            return

        initial_dir = state_path_candidate(self.gui_state.get("last_frequency_config_dir"))
        if initial_dir is None:
            initial_dir = DEFAULT_DOCUMENTS_BROWSE_DIR
        path = filedialog.asksaveasfilename(
            parent=self,
            title="Save frequency config",
            initialdir=str(initial_dir),
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
        )
        if not path:
            return
        config_path = Path(path)
        try:
            write_json(config_path, self.frequency_config_payload())
        except Exception as exc:
            messagebox.showerror("Save Config", str(exc), parent=self)
            return
        self.gui_state["last_frequency_config_dir"] = str(config_path.parent)
        write_gui_state(self.gui_state)
        self.status.set(f"Frequency config saved: {config_path.name}")
        messagebox.showinfo("Save Config", f"Saved:\n{config_path}", parent=self)

    def recall_frequency_config(self) -> None:
        initial_dir = state_path_candidate(self.gui_state.get("last_frequency_config_dir"))
        if initial_dir is None:
            initial_dir = DEFAULT_DOCUMENTS_BROWSE_DIR
        path = filedialog.askopenfilename(
            parent=self,
            title="Recall frequency config",
            initialdir=str(initial_dir),
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
        )
        if not path:
            return
        config_path = Path(path)
        try:
            payload = read_json(config_path)
            self.apply_frequency_config_payload(payload)
        except Exception as exc:
            messagebox.showerror("Recall Config", str(exc), parent=self)
            return
        self.gui_state["last_frequency_config_dir"] = str(config_path.parent)
        write_gui_state(self.gui_state)
        self.status.set(f"Frequency config recalled: {config_path.name}")
        messagebox.showinfo("Recall Config", f"Loaded:\n{config_path}", parent=self)

    def apply_frequency_config_payload(self, payload: dict[str, Any]) -> None:
        version = payload.get("version")
        if version != f"easy-frequency-config-{APP_VERSION}":
            raise ValueError("Unsupported frequency config version.")

        default_parameters = payload.get("default_parameters")
        if not isinstance(default_parameters, dict):
            raise ValueError("default_parameters must be a JSON object.")

        for label, _parameters in PARAMETER_ROWS:
            raw = default_parameters.get(label)
            if not isinstance(raw, dict):
                continue
            self.parameter_include_vars[label].set(bool(raw.get("include", False)))
            frequency_text = str(raw.get("frequency_points", "")).strip()
            split_frequency_text(frequency_text)
            self.parameter_frequency_vars[label].set(frequency_text)
            plot_limits = raw.get("plot_limits")
            if isinstance(plot_limits, dict):
                self.parameter_plot_limits[label] = copy.deepcopy(plot_limits)
            else:
                self.parameter_plot_limits.pop(label, None)

        additional_parameters = payload.get("additional_parameters", [])
        if not isinstance(additional_parameters, list):
            raise ValueError("additional_parameters must be a JSON list.")

        rebuilt_rows: list[dict[str, Any]] = []
        for index, raw in enumerate(additional_parameters, start=1):
            if not isinstance(raw, dict):
                raise ValueError(f"additional_parameters[{index}] must be a JSON object.")
            source = str(raw.get("source", "MAIN")).strip().upper()
            if source not in FILE_GROUPS:
                raise ValueError(f"additional_parameters[{index}] has invalid source.")
            parameter = batch_core.normalize_s_parameter(str(raw.get("parameter", "")).strip())
            if not parameter or not batch_core.is_supported_parameter(parameter):
                raise ValueError(f"additional_parameters[{index}] has unsupported parameter.")
            frequency_text = str(raw.get("frequency_points", "")).strip()
            split_frequency_text(frequency_text)
            plot_limits = raw.get("plot_limits")
            rebuilt_rows.append(
                {
                    "source": source,
                    "parameter": parameter,
                    "frequency_points": frequency_text,
                    "plot_limits": copy.deepcopy(plot_limits) if isinstance(plot_limits, dict) else None,
                }
            )

        selected_frequency_map(self.parameter_settings())
        self.clear_additional_rows()
        for row in rebuilt_rows:
            self.add_additional_row(
                row["source"],
                row["parameter"],
                row["frequency_points"],
                row["plot_limits"],
            )
        self.additional_targets_for_config()

    def default_plot_limits_for_label(self, label: str) -> dict[str, Any]:
        plot_key = PLOT_KEY_BY_LABEL[label]
        defaults = copy.deepcopy(SUPPLIER_DEFAULT_PLOT_LIMITS)
        if plot_key in SUPPLIER_PLOT_Y_LIMITS:
            defaults["y_db"] = copy.deepcopy(SUPPLIER_PLOT_Y_LIMITS[plot_key])
        return merge_plot_limit_defaults(None, defaults)

    def edit_default_plot_setting(self, label: str) -> None:
        self.open_plot_setting_dialog(
            title=f"Plot Setting - {label}",
            initial_limits=self.parameter_plot_limits.get(label),
            default_limits=self.default_plot_limits_for_label(label),
            on_save=lambda limits: self.save_default_plot_setting(label, limits),
        )

    def save_default_plot_setting(self, label: str, limits: dict[str, Any]) -> None:
        self.parameter_plot_limits[label] = limits
        self.status.set(f"Plot setting updated: {label}")

    def add_additional_row(
        self,
        source: str = "MAIN",
        parameter: str = "S11",
        frequency_points: str = "100MHz",
        plot_limits: dict[str, Any] | None = None,
    ) -> None:
        row_id = uuid.uuid4().hex
        frame = ttk.Frame(self.additional_rows_frame, style="Card.TFrame")
        frame.pack(fill="x", pady=4)
        source_var = tk.StringVar(value=source if source in FILE_GROUPS else "MAIN")
        parameter_var = tk.StringVar(value=parameter)
        frequency_var = tk.StringVar(value=frequency_points)
        ttk.Combobox(
            frame,
            textvariable=source_var,
            values=FILE_GROUPS,
            state="readonly",
            width=10,
        ).grid(row=0, column=0, sticky="w", padx=(0, 8))
        ttk.Entry(frame, textvariable=parameter_var, width=16).grid(
            row=0,
            column=1,
            sticky="w",
            padx=(0, 8),
        )
        ttk.Entry(frame, textvariable=frequency_var).grid(
            row=0,
            column=2,
            sticky="ew",
            padx=(0, 8),
        )
        ttk.Button(
            frame,
            text="Plot Setting",
            command=lambda current_id=row_id: self.edit_additional_row_plot_setting(current_id),
            style="Tool.TButton",
        ).grid(row=0, column=3, sticky="w", padx=(0, 6))
        ttk.Button(
            frame,
            text="Remove",
            command=lambda current_id=row_id: self.remove_additional_row(current_id),
            style="Tool.TButton",
        ).grid(row=0, column=4, sticky="w")
        frame.columnconfigure(2, weight=1)
        self.additional_rows[row_id] = {
            "frame": frame,
            "source_var": source_var,
            "parameter_var": parameter_var,
            "frequency_var": frequency_var,
            "plot_limits": copy.deepcopy(plot_limits) if isinstance(plot_limits, dict) else default_additional_plot_limits(),
        }

    def edit_additional_row_plot_setting(self, row_id: str) -> None:
        row = self.additional_rows.get(row_id)
        if row is None:
            return
        source = row["source_var"].get().strip().upper() or "MAIN"
        parameter = row["parameter_var"].get().strip() or "S-parameter"
        self.open_plot_setting_dialog(
            title=f"Plot Setting - {source} {parameter}",
            initial_limits=row.get("plot_limits"),
            default_limits=default_additional_plot_limits(),
            on_save=lambda limits: self.save_additional_row_plot_setting(row_id, limits),
        )

    def save_additional_row_plot_setting(self, row_id: str, limits: dict[str, Any]) -> None:
        if row_id in self.additional_rows:
            self.additional_rows[row_id]["plot_limits"] = limits
            self.status.set("Additional plot setting updated.")

    def remove_additional_row(self, row_id: str) -> None:
        row = self.additional_rows.pop(row_id, None)
        if row is not None:
            row["frame"].destroy()

    def clear_additional_rows(self) -> None:
        for row_id in list(self.additional_rows):
            self.remove_additional_row(row_id)

    def open_plot_setting_dialog(
        self,
        title: str,
        initial_limits: dict[str, Any] | None,
        default_limits: dict[str, Any],
        on_save,
    ) -> None:
        limits = merge_plot_limit_defaults(initial_limits, default_limits)
        window = tk.Toplevel(self)
        window.title(title)
        window.resizable(False, False)
        window.transient(self)
        window.configure(bg=THEME["app_bg"])
        frame = ttk.Frame(window, padding=14, style="App.TFrame")
        frame.pack(fill="both", expand=True)

        x_frequency = limits.get("x_frequency", {})
        y_db = limits.get("y_db", {})
        x_min_var = tk.StringVar(value=str(x_frequency.get("min") or ""))
        x_max_var = tk.StringVar(value=str(x_frequency.get("max") or ""))
        y_min_var = tk.StringVar(value=str(y_db.get("min") if y_db.get("min") is not None else ""))
        y_max_var = tk.StringVar(value=str(y_db.get("max") if y_db.get("max") is not None else ""))

        fields = (
            ("X min frequency", x_min_var, "Example: 1MHz"),
            ("X max frequency", x_max_var, "Example: 500MHz"),
            ("Y min dB", y_min_var, "Example: -50"),
            ("Y max dB", y_max_var, "Example: 0"),
        )
        for row_index, (label, variable, example) in enumerate(fields):
            ttk.Label(frame, text=label).grid(row=row_index, column=0, sticky="w", pady=4)
            ttk.Entry(frame, textvariable=variable, width=24).grid(
                row=row_index,
                column=1,
                sticky="ew",
                padx=(8, 8),
                pady=4,
            )
            ttk.Label(frame, text=example, foreground=THEME["muted"]).grid(
                row=row_index,
                column=2,
                sticky="w",
                pady=4,
            )

        def save() -> None:
            try:
                new_limits = build_plot_limits_from_text(
                    x_min_var.get(),
                    x_max_var.get(),
                    y_min_var.get(),
                    y_max_var.get(),
                )
            except Exception as exc:
                messagebox.showerror("Plot Setting", str(exc), parent=window)
                return
            on_save(new_limits)
            window.destroy()

        buttons = ttk.Frame(frame)
        buttons.grid(row=len(fields), column=0, columnspan=3, sticky="e", pady=(10, 0))
        ttk.Button(buttons, text="Save", command=save).pack(side="left", padx=(0, 6))
        ttk.Button(buttons, text="Cancel", command=window.destroy).pack(side="left")
        self.center_child_window(window)

    def choose_output(self) -> None:
        last_output = state_path_candidate(self.gui_state.get("last_output_dir"))
        current_output = current_entry_dir(self.output_dir.get())
        selected = choose_folder_dialog(
            title="Choose output folder",
            initial_dir=existing_initial_dir(
                *(
                    candidate
                    for candidate in (last_output, current_output, DEFAULT_DOCUMENTS_BROWSE_DIR, APP_DIR)
                    if candidate is not None
                )
            ),
        )
        if selected:
            self.output_dir.set(selected)
            self.gui_state["last_output_dir"] = selected
            write_gui_state(self.gui_state)

    def add_s4p_files(self) -> None:
        group = self.selected_group.get().strip().upper()
        if group not in FILE_GROUPS:
            messagebox.showerror("Add files", "Choose MAIN, NEXT, or FEXT first.", parent=self)
            return
        last_file_dir = state_path_candidate(self.gui_state.get("last_s4p_file_dir"))
        paths = filedialog.askopenfilenames(
            title=f"Choose {group} .s4p files",
            initialdir=existing_initial_dir(
                *(
                    candidate
                    for candidate in (last_file_dir, DEFAULT_DOCUMENTS_BROWSE_DIR, APP_DIR)
                    if candidate is not None
                )
            ),
            filetypes=[("Touchstone 4-port", "*.s4p"), ("All files", "*.*")],
        )
        if not paths:
            return
        added = 0
        for raw_path in paths:
            path = Path(raw_path)
            if path.suffix.lower() != ".s4p":
                continue
            if self.add_file_to_table(group, path):
                added += 1
        if paths:
            self.gui_state["last_s4p_file_dir"] = str(Path(paths[-1]).parent)
            write_gui_state(self.gui_state)
        self.refresh_file_summary()
        self.suggest_output_name()
        self.output_status.set(f"Added {added} {group} file(s).")

    def add_file_to_table(self, group: str, path: Path) -> bool:
        normalized_path = str(path.resolve())
        for item in self.selected_files.values():
            if item["group"] == group and item["path"] == normalized_path:
                return False
        row_id = uuid.uuid4().hex
        self.selected_files[row_id] = {
            "group": group,
            "file_name": path.name,
            "path": normalized_path,
        }
        self.file_table.insert(
            "",
            "end",
            iid=row_id,
            values=(group, path.name, normalized_path),
        )
        return True

    def remove_selected_files(self) -> None:
        for row_id in self.file_table.selection():
            self.selected_files.pop(str(row_id), None)
            self.file_table.delete(row_id)
        self.refresh_file_summary()

    def clear_selected_files(self) -> None:
        self.selected_files.clear()
        for row_id in self.file_table.get_children():
            self.file_table.delete(row_id)
        self.refresh_file_summary()

    def selected_file_counts(self) -> dict[str, int]:
        counts = {group: 0 for group in FILE_GROUPS}
        for item in self.selected_files.values():
            counts[item["group"]] += 1
        return counts

    def refresh_file_summary(self) -> None:
        counts = self.selected_file_counts()
        self.file_count_status.set(
            f"Selected files: MAIN {counts['MAIN']}, NEXT {counts['NEXT']}, FEXT {counts['FEXT']}."
        )

    def suggest_output_name(self) -> None:
        if self.custom_output_name.get() or self.output_name.get().strip():
            return
        for item in self.selected_files.values():
            if item["group"] == "MAIN":
                suggested = Path(item["file_name"]).stem
                try:
                    validate_output_file_name(suggested)
                except ValueError:
                    suggested = "supplier_output"
                self.output_name.set(suggested)
                return
        if self.selected_files:
            self.output_name.set("supplier_output")

    def prepare_supplier_input_dir(self) -> Path:
        temp_root = Path(tempfile.gettempdir()) / "xfmr_supplier_simple"
        temp_root.mkdir(parents=True, exist_ok=True)
        run_dir = temp_root / f"run_{uuid.uuid4().hex}"
        input_dir = run_dir / "input"
        xtk_dir = input_dir / "XTK"
        input_dir.mkdir(parents=True, exist_ok=True)
        xtk_dir.mkdir(parents=True, exist_ok=True)

        group_indices = {group: 0 for group in FILE_GROUPS}
        for item in self.selected_files.values():
            group = item["group"]
            group_indices[group] += 1
            source = Path(item["path"])
            if group == "MAIN":
                destination = input_dir / f"supplier_MAIN_{group_indices[group]:04d}.s4p"
            else:
                destination = xtk_dir / (
                    f"supplier_{group}_{group_indices[group]:04d}_{group}.s4p"
                )
            shutil.copy2(source, destination)
        self.run_temp_dir = run_dir
        return input_dir

    def cleanup_temp_run_dir(self) -> None:
        if self.run_temp_dir is None:
            return
        temp_root = Path(tempfile.gettempdir()) / "xfmr_supplier_simple"
        try:
            resolved_run = self.run_temp_dir.resolve()
            resolved_root = temp_root.resolve()
            if str(resolved_run).startswith(str(resolved_root)):
                shutil.rmtree(resolved_run, ignore_errors=True)
        finally:
            self.run_temp_dir = None

    def _toggle_output_mode(self) -> None:
        if self.custom_output_name.get():
            self.output_combo.configure(state="normal")
            self.output_status.set("Custom output file name is enabled.")
        else:
            self.output_combo.configure(state="readonly")
            self.output_status.set("Output file name can be auto-suggested from MAIN files.")

    def validate_settings(self) -> tuple[Path, str, str, bool]:
        output_text = self.output_dir.get().strip()
        output_name = self.output_name.get().strip()
        custom_name = bool(self.custom_output_name.get())
        counts = self.selected_file_counts()
        if counts["MAIN"] < 1:
            raise ValueError("Add at least one MAIN .s4p file.")
        for item in self.selected_files.values():
            path = Path(item["path"])
            if not path.is_file():
                raise ValueError(f"Selected file does not exist: {path}")
        if not output_text:
            raise ValueError("Output folder is required.")
        if not output_name:
            raise ValueError("Output file name is required.")
        output_name = validate_output_file_name(output_name)
        selected_frequency_map(self.parameter_settings())
        input_dir = self.prepare_supplier_input_dir()
        return input_dir, output_text, output_name, custom_name

    def plot_limit_overrides_for_config(self) -> dict[str, dict[str, Any]]:
        overrides: dict[str, dict[str, Any]] = {}
        for label, limits in self.parameter_plot_limits.items():
            plot_key = PLOT_KEY_BY_LABEL.get(label)
            if plot_key:
                overrides[plot_key] = copy.deepcopy(limits)
        return overrides

    def additional_targets_for_config(self) -> list[dict[str, Any]]:
        targets = []
        seen: set[tuple[str, str]] = set()
        for row in self.additional_rows.values():
            source = row["source_var"].get().strip().upper()
            parameter_text = row["parameter_var"].get().strip()
            frequency_text = row["frequency_var"].get().strip()
            if not parameter_text and not frequency_text:
                continue
            if source not in FILE_GROUPS:
                raise ValueError("Additional Parameters contains an invalid Source.")
            parameter = batch_core.normalize_s_parameter(parameter_text)
            if not parameter:
                raise ValueError("Additional Parameters contains a blank Parameter.")
            if not batch_core.is_supported_parameter(parameter):
                raise ValueError(
                    "Additional Parameters contains an unsupported parameter. "
                    "Examples: S11, S41, Sdd21, Scd22, Sdc11, Scc21."
                )
            frequencies = split_frequency_text(frequency_text)
            if not frequencies:
                raise ValueError(f"Additional Parameters {source} {parameter} has no frequency points.")
            key = (source, parameter)
            if key in seen:
                raise ValueError(f"Additional Parameters has duplicate row: {source} {parameter}.")
            seen.add(key)
            targets.append(
                {
                    "include": True,
                    "groups": [source],
                    "parameter": parameter,
                    "frequencies": frequencies,
                    "plot_limits": copy.deepcopy(row.get("plot_limits")),
                }
            )
        return targets

    def start_processing(self) -> None:
        if self.process is not None:
            messagebox.showwarning("Processing", "Processing is already running.")
            return
        try:
            input_path, output_text, output_name, custom_name = self.validate_settings()
            if self.run_temp_dir is None:
                raise ValueError("Temporary input folder was not prepared.")
            config = build_supplier_config(
                str(input_path),
                output_text,
                output_name,
                custom_name,
                self.parameter_settings(),
                self.plot_limit_overrides_for_config(),
                self.additional_targets_for_config(),
                self.run_temp_dir,
            )
        except Exception as exc:
            self.cleanup_temp_run_dir()
            messagebox.showerror("Settings error", str(exc), parent=self)
            return

        if self.run_temp_dir is None:
            messagebox.showerror("Settings error", "Temporary input folder was not prepared.", parent=self)
            return
        config_path = self.run_temp_dir / f"supplier_config_{uuid.uuid4().hex}.json"
        write_json(config_path, config)

        self.gui_state["last_output_dir"] = output_text
        write_gui_state(self.gui_state)
        self.completed_excel_path = None
        self.progress_value.set(0)
        self.progress_label.set("0%")
        self.log_text.delete("1.0", "end")
        self.status.set("Processing...")
        self.start_button.configure(state="disabled")
        self.stop_button.configure(state="normal")

        command = self.batch_command(config_path)
        self.process = subprocess.Popen(
            command,
            cwd=str(APP_DIR),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
        threading.Thread(target=self._read_process_output, args=(self.process,), daemon=True).start()

    def batch_command(self, config_path: Path) -> list[str]:
        if getattr(sys, "frozen", False):
            return [sys.executable, "--easy-batch-main", "--config", str(config_path)]
        return [sys.executable, str(APP_DIR / "EASYmain.py"), "--config", str(config_path)]

    def _read_process_output(self, process: subprocess.Popen[str]) -> None:
        if process.stdout is None:
            return
        for line in process.stdout:
            self.process_queue.put(line)
        process.wait()
        self.process_queue.put(f"__PROCESS_EXIT__:{process.returncode}")

    def _poll_process_queue(self) -> None:
        try:
            while True:
                line = self.process_queue.get_nowait()
                if line.startswith("__PROCESS_EXIT__:"):
                    return_code = int(line.split(":", 1)[1])
                    self._finish_process(return_code)
                    continue
                self._append_log_line(line)
        except queue.Empty:
            pass
        self.after(120, self._poll_process_queue)

    def _append_log_line(self, line: str) -> None:
        self.log_text.insert("end", line)
        self.log_text.see("end")
        progress_match = re.search(r"PROGRESS:\s*(\d+)", line)
        if progress_match:
            value = max(0, min(100, int(progress_match.group(1))))
            self.progress_value.set(value)
            self.progress_label.set(f"{value}%")
        excel_match = re.search(r"Wrote Excel:\s*(.+)$", line.strip())
        if excel_match:
            self.completed_excel_path = Path(excel_match.group(1).strip())

    def _finish_process(self, return_code: int) -> None:
        self.process = None
        self.start_button.configure(state="normal")
        self.stop_button.configure(state="disabled")
        self.cleanup_temp_run_dir()
        if return_code == 0:
            self.progress_value.set(100)
            self.progress_label.set("100%")
            try:
                if self.completed_excel_path is not None and self.completed_excel_path.exists():
                    trim_supplier_workbook(self.completed_excel_path)
                    self._append_log_line(
                        "Supplier workbook trimmed to sheets: "
                        f"{', '.join(SUPPLIER_EXCEL_SHEETS)}\n"
                    )
            except Exception as exc:
                self.status.set("Error.")
                messagebox.showerror(
                    "Excel trim error",
                    f"Processing finished, but supplier Excel cleanup failed:\n{exc}",
                    parent=self,
                )
                return
            self.status.set("Completed.")
            self.show_completion_dialog()
        elif return_code < 0:
            self.status.set("Stopped.")
            messagebox.showinfo("Processing stopped", "Processing was stopped by user.", parent=self)
        else:
            self.status.set("Error.")
            messagebox.showwarning(
                "Processing finished",
                "Processing reported a warning or error. Please review the log.",
                parent=self,
            )

    def stop_processing(self) -> None:
        if self.process is None:
            return
        process = self.process
        self.status.set("Stopping...")
        self.stop_button.configure(state="disabled")
        try:
            process.terminate()
        except OSError:
            return
        self.after(2000, lambda: self._kill_if_running(process))

    def _kill_if_running(self, process: subprocess.Popen[str]) -> None:
        if process.poll() is None:
            try:
                process.kill()
            except OSError:
                pass

    def show_completion_dialog(self) -> None:
        window = tk.Toplevel(self)
        window.title("Processing complete")
        window.resizable(False, False)
        window.transient(self)
        frame = ttk.Frame(window, padding=16)
        frame.pack(fill="both", expand=True)
        excel_text = str(self.completed_excel_path) if self.completed_excel_path else "not found"
        ttk.Label(
            frame,
            text=f"Processing complete\nOutput file name: {self.output_name.get().strip()}\nExcel: {excel_text}",
            justify="left",
        ).pack(fill="x", pady=(0, 14))
        buttons = ttk.Frame(frame)
        buttons.pack(fill="x")
        open_button = ttk.Button(buttons, text="Open Excel", command=lambda: self.open_excel(window))
        open_button.pack(side="left")
        if self.completed_excel_path is None or not self.completed_excel_path.exists():
            open_button.configure(state="disabled")
        ttk.Button(buttons, text="OK", command=window.destroy).pack(side="right")
        self.center_child_window(window)
        window.grab_set()

    def center_child_window(self, window: tk.Toplevel) -> None:
        self.update_idletasks()
        window.update_idletasks()
        x = self.winfo_rootx() + max((self.winfo_width() - window.winfo_reqwidth()) // 2, 0)
        y = self.winfo_rooty() + max((self.winfo_height() - window.winfo_reqheight()) // 2, 0)
        window.geometry(f"+{x}+{y}")

    def open_excel(self, window: tk.Toplevel | None = None) -> None:
        if self.completed_excel_path is None or not self.completed_excel_path.exists():
            messagebox.showerror("Open Excel", "Excel file was not found.", parent=self)
            return
        try:
            os.startfile(self.completed_excel_path)  # type: ignore[attr-defined]
        except OSError as exc:
            messagebox.showerror("Open Excel", str(exc), parent=self)
            return
        if window is not None:
            window.destroy()


def run_batch_main() -> int:
    import EASYmain

    sys.argv = [
        sys.argv[0],
        *(arg for arg in sys.argv[1:] if arg != "--easy-batch-main"),
    ]
    return EASYmain.main()


def main() -> None:
    app = SupplierSimpleGui()
    app.mainloop()


if __name__ == "__main__":
    if "--easy-batch-main" in sys.argv:
        raise SystemExit(run_batch_main())
    main()
